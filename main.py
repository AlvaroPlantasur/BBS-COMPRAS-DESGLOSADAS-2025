import os
import pandas as pd
import psycopg2
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def conectar_db():
    return psycopg2.connect(
        dbname=os.environ["DB_NAME"],
        user=os.environ["DB_USER"],
        password=os.environ["DB_PASSWORD"],
        host=os.environ["DB_HOST"],
        port=os.environ["DB_PORT"]
    )

def ejecutar_consulta(conn):
    query = f"""
        select 
        ail.invoice_id as "ID FACTURA",
        ai.name as "REFERENCIA ALBARÁN",
        ai.internal_number as "CÓDIGO FACTURA",
        to_char(ai.date_invoice, 'DD/MM/YYYY') as "FECHA FACTURA",
        pp.default_code as "REFERENCIA PRODUCTO", 
        pp.name as "NOMBRE", 
        s.name AS "SECCION", 
        f.name as "FAMILIA", 
        sf.name as "SUBFAMILIA",
        rc.name as "COMPAÑÍA",
        ssp.name as "SEDE",
        (CASE WHEN stp.directo_cliente = true THEN 'Sí' ELSE 'No' END) AS "CAMIÓN DIRECTO",
        SUM(stp.cargos_extra_prorrateo) AS "CARGOS EXTRA",
        ai.portes as "PORTES",
        rp.nombre_comercial as "CLIENTE",
        rp.vat as "CIF CLIENTE",
        c.name as "PAÍS",
        extract(MONTH FROM ai.date_invoice) as "MES",
        extract(MONTH FROM ai.date_invoice) as "DÍA",
        (
            case when ai.type = 'in_invoice' then ail.cantidad_pedida
            when ai.type = 'in_refund' then -ail.cantidad_pedida
            end
        ) as "UNIDADES COMPRA",
        (
            case when ai.type = 'in_invoice' then ail.price_subtotal
            when ai.type = 'in_refund' then -ail.price_subtotal
            end
        ) as "BASE COMPRA TOTAL",
        (
            case when ai.type = 'in_invoice' then sum(case when coalesce(at.amount,0) = 1 then 0.0 else (coalesce(at.amount,0)*ail.price_subtotal) end)
            when ai.type = 'in_refund' then -sum(case when coalesce(at.amount,0) = 1 then 0.0 else (coalesce(at.amount,0)*ail.price_subtotal) end)
            end
        ) as "IMPUESTOS",
        (
            case when ai.type = 'in_invoice' then ail.price_subtotal + sum(case when coalesce(at.amount,0) = 1 then 0.0 else (coalesce(at.amount,0)*ail.price_subtotal) end)
            when ai.type = 'in_refund' then -(ail.price_subtotal + sum(case when coalesce(at.amount,0) = 1 then 0.0 else (coalesce(at.amount,0)*ail.price_subtotal) end))
            end
        ) as "IMPORTE COMPRA TOTAL"
    
    from account_invoice_line ail
    inner join account_invoice ai ON ai.id = ail.invoice_id
    inner join product_product pp ON ail.product_id = pp.id
    inner join res_partner rp on rp.id = ai.partner_id
    inner join res_partner_address rpa ON rpa.id = ai.address_invoice_id
    inner join res_country c on c.id = rpa.pais_id
    left outer join stock_picking stp ON stp.name = split_part(ai.origin,':', 1)
    left outer join res_company rc on rc.id = ai.company_id
    left outer join stock_sede_ps ssp on ssp.id = ai.sede_id
    left outer join product_category s ON (s.id = pp.seccion)
    left outer join product_category f ON (f.id = pp.familia)
    left outer join product_category sf ON (sf.id = pp.subfamilia)
    left outer join account_invoice_line_tax ailt on ail.id = ailt.invoice_line_id
    left outer join account_tax at on ailt.tax_id = at.id
    where ai.state in ('open','paid') and ai.type in ('in_invoice','in_refund') and ai.date_invoice >= '{fecha_inicio_str}' and ai.date_invoice <= '{fecha_fin_str}' and ai.obsolescencia = false
    group by 
        ail.id,
        rp.id,
        ail.company_id,
        ai.sede_id,
        ai.date_invoice,
        to_char(ai.date_invoice, 'YYYY'),
        to_char(ai.date_invoice, 'MM'),
        to_char(ai.date_invoice, 'YYYY-MM-DD'),
        pp.seccion,
        pp.familia,
        pp.subfamilia,
        pp.default_code,
        pp.id,
        ai.partner_id,
        ai.anticipo,
        c.name,
        rpa.prov_id,
        rpa.state_id_2,
        ai.name,
        ai.internal_number,
        ai.origin,
        ail.cantidad_pedida,
        ail.price_subtotal,
        s.name,
        f.name,
        sf.name,
        rc.name,
        ssp.name,
        ai.directo_cliente,
        ai.portes,
        rp.nombre_comercial,
        ai.type,
        stp.directo_cliente,
        rp.vat;
    """
    return pd.read_sql_query(query, conn)

def actualizar_excel(df, file_path, sheet_name="CompDesglosadas2025"):
    print(f"Se obtuvieron {len(df)} filas de la consulta.")

    try:
        book = load_workbook(file_path)

        if sheet_name not in book.sheetnames:
            raise ValueError(f"La hoja '{sheet_name}' no existe en el archivo Excel.")

        sheet = book[sheet_name]

        # Borrar todas las filas excepto la cabecera
        sheet.delete_rows(2, sheet.max_row)

        # Escribir el DataFrame (sin headers ni index)
        for row in dataframe_to_rows(df, index=False, header=False):
            sheet.append(row)

        book.save(file_path)
        print("✅ Datos actualizados correctamente en la hoja", sheet_name)

    except Exception as e:
        print("❌ Error al actualizar el archivo Excel:", str(e))
        raise

def main():
    try:
        conn = conectar_db()
        df = ejecutar_consulta(conn)
        conn.close()

        file_path = os.environ.get("EXCEL_FILE_PATH")
        actualizar_excel(df, file_path)

    except Exception as e:
        print("❌ Error general:", str(e))
        raise

if __name__ == "__main__":
    main()
